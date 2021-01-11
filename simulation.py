import random
import numpy as np
import time
import codecs
import json
import os
import multiprocessing
import xlwt
from openpyxl import load_workbook

#Creates a bingo board with random numbers following bingo rules
def createBoards(numOfBoards):
    #Creates a bingo board
    #parameters: numOfBoards - How many boards to createBoards
    #return: a numpy array containing all the bingo boards created, which are 5x5 numpy arrays
    bingoCards = (np.empty((int(numOfBoards), 5, 5)))
    for i in range(numOfBoards):
        board = np.empty((5,5))
        for j in range(5):
            board[j]= random.sample(range((15 * j) + 1, (15 * (j+1)) + 1 ), 5)
        board = np.transpose(board, (1,0))
        bingoCards[i] = board
    return bingoCards


#Runs a bingo game
def simulateGame(numOfGames, numOfBoards):
    #Simulates a bingo game
    #parameters: numOfGames - How many games to simulates
    #            numOfBoards - How many boards are in the game
    #return: adds a json file to the tmp folder containing the number of horizontal wins and vertical wins achieved each round
    #as well as how many new horizontal and vertical wins acheived each round
    totalHWins = np.empty((75))
    totalVWins = np.empty((75))
    totalHWins.fill(0)
    totalVWins.fill(0)
    newStartTime = time.time()
    newEndTime = 0
    for i in range(0, numOfGames):
        #creates a random sequence of 75 numbers which will simulate every bingo call
        calls = random.sample(range(1,76), 75)
        #creates a number of boards designated by the function paramters
        bC = createBoards(numOfBoards)
        #Prints a statement telling how many games have been completed 10 times every total game
        if (i%(numOfGames/10)==0 and i!=0):
            newEndTime = time.time()
            totalTime = (newEndTime-newStartTime)/60
            print("Number of Players: " + str(numOfBoards) + " - " + str(i) + " Completed - " + str(totalTime) + " minutes")
            newStartTime = time.time()

        for j in range (0, 75):
            #For every number that is called, the value of that number in the bingo board is changed to 0
            #using this, if the sum of a bingo row/column is 0, then you know it's a bingo
            bC[bC == calls[j]] = 0
            hWin = np.int(np.count_nonzero(bC.sum(axis=2) == 0))
            vWin = np.int(np.count_nonzero(bC.sum(axis=1) == 0))
            #once number of bingo wins is >0, it increments one to the bingo counter, as a horizontal win or vertical win was achieved on that turn
            if (hWin > 0):
                totalHWins[j] += 1
            if (vWin > 0):
                totalVWins[j] += 1
            #The program then exits the for loop and starts the next game
            if (hWin > 0 or vWin > 0):
                break

    print("Number of Players: " + str(numOfBoards) + " - Complete")
    return totalHWins, totalVWins

def makeJson(arrayOfTotalHWins, arrayOfTotalVWins):
    jsonArray = np.empty((1 + 2*(len(arrayOfTotalHWins)), 75))
    #creates a counting array to number each turn that occurs
    jsonArray[0] = np.arange(1, 76, dtype=np.int)

    for j in range (0, len(arrayOfTotalHWins)):
        jsonArrayNum = j*2
        jsonArray[jsonArrayNum+1] = arrayOfTotalHWins[j]
        jsonArray[jsonArrayNum+2] = arrayOfTotalVWins[j]

    jsonArray = jsonArray.transpose()
    args = (jsonArray)
    dataArray = np.vstack(args).tolist()

    i = 1
    while os.path.exists("tmp/Bingo_Data_%s.json" % i):
        i += 1

    filePath = "tmp/Bingo_Data_" + str(i) + ".json"

    json.dump(dataArray, codecs.open(filePath, 'w', encoding='utf-8'), separators=(',', ':'), sort_keys=True, indent=4) ### this saves the array in .json format
    return filePath

def makeExcel(filePath, gameVals):
    wb = load_workbook("Bingo_Data_Template.xlsx")
    sheet1 = wb["Sheet 1"]

    with open(filePath) as json_file:
        data = json.load(json_file)
        cell = sheet1.cell(row=1, column=1)
        cell.value = "Call Number"

        for i in range(1, len(data[0])):
            gameValsIndexNum = int((i-1)/2)
            cell = sheet1.cell(row = 1, column = i+1)
            if(i%2 != 0):
                cell.value = str(gameVals[gameValsIndexNum][0]) + " Game(s), " + str(gameVals[gameValsIndexNum][1]) + " Player(s): Horizontal Wins"
            elif(i%2 == 0):
                cell.value = str(gameVals[gameValsIndexNum][0]) + " Game(s), " + str(gameVals[gameValsIndexNum][1]) + " Player(s): Vertical Wins"
        for i in range (1, 76):
            for j in range (0, len(data[0])):
                cell = sheet1.cell(row = i+1, column = j+1)
                cell.value = data[i-1][j]

    newFilePath = filePath.replace(".json",".xlsx")
    wb.save(newFilePath)

def collectGameData(gameVals):
    totalHorizontalWins = np.empty((len(gameVals), 75))
    totalVerticalWins = np.empty((len(gameVals), 75))

    for i in range (0, len(gameVals)):
        hWins, vWins = simulateGame(gameVals[i][0], gameVals[i][1])
        totalHorizontalWins[i] = hWins
        totalVerticalWins[i] = vWins
    makeExcel(makeJson(totalHorizontalWins, totalVerticalWins), gameVals)



# Type in game vals below
# First value is number of games
# Second value is number of players
start = time.time()
gameVals = [
[100000, 1],
[100000, 10],
[100000, 100],
[100000, 1000],
[100000, 10000]
]

collectGameData(gameVals)

end = time.time()
print((end - start)/60)
